# Copyright (c) 2022, GreyCube Technologies and contributors
# For license information, please see license.txt

import frappe
from frappe import _
from frappe.utils import now, cstr, cint, add_days, today, add_to_date
import json
from art_collections.controllers.excel import write_xlsx, attach_file, add_images
import openpyxl
from openpyxl.styles.alignment import Alignment
import io
import os
from openpyxl import load_workbook
from erpnext.accounts.party import get_party_details
import re
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.cell import get_column_letter


LEAD_ITEM_FIELDS = """
'item_photo',
'supplier_part_no',
'supplier_item_description',
'product_material1',
'percentage1',
'product_material2',
'percentage2',
'product_material3',
'percentage3',
'customs_tariff_number',
'item_length',
'item_width',
'item_thickness',
'packing_type',
'racking_bag',
'minimum_order_qty',
'unit_price',
'item_price',
'item_price_valid_from',
'pricing_rule_qty',
'pricing_rule_price',
'pricing_rule_valid_from',
'inner_qty',
'is_display_box',
'is_special_mentions',
'is_certificates_reqd',
'test_report',
'lead_item_name',
'item_group',
'description',
'uom',
'selling_pack_qty',
'is_disabled',
'is_quoted',
'is_sample_validated',
'is_item_created',
'is_po_created',
'designation',
'description1',
'description2',
'description3',
'other_language',
'is_need_photo_for_packaging',
'packaging_description_excel',
'name',
'photo_quotation'
"""


def get_lead_item_fields():

    fields = frappe.db.sql(
        """
			select fieldname , label , 
			case 
				when fieldtype = 'Attach Image' then 'text'
				when fieldtype in ('Percent', 'Int', 'Currency') then 'numeric' 
				when fieldtype in ('Check') then 'checkbox'
				when fieldtype in ('Date') then 'calendar'
			else 'text' end fieldtype 
			from tabDocField tdf where parent = 'Lead Item' 
			and label is not null
			and fieldname in ({fields})
			ORDER BY FIELD (fieldname,{fields})
	""".format(
            fields=LEAD_ITEM_FIELDS
        ),
        as_dict=True,
        # debug=True,
    )

    return fields + [
        frappe._dict(
            {
                "fieldname": "name",
                "label": "Lead Item#",
                "fieldtype": "text",
            }
        )
    ]


def get_items_xlsx(docname, template="", supplier=None, filters=None):
    SHEET_NAME = "Lead Items"

    template_file_path = os.path.join(os.path.dirname(__file__), template + ".xlsx")
    if not os.path.exists(template_file_path):
        frappe.throw("No xlsx template found for " + template)

    wb = load_workbook(template_file_path)
    ws = wb["configuration"]
    fields = [col.value for col in ws["B"]][1:]
    skip_rows = ws["C2"].value

    # print(ws.max_row, ws.max_column, fields, skip_rows)

    data = frappe.db.sql(
        """
    	select {} from `tabLead Item` where photo_quotation = %s {}""".format(
            ", ".join(fields), LEAD_ITEM_CONDITIONS.get(filters, "")
        ),
        (docname,),
        as_list=1,
    )

    photo_index = "item_photo" in fields and fields.index("item_photo") or 0

    excel_rows = list(data)
    images = [d[photo_index] for d in data]

    wb = write_xlsx(
        excel_rows,
        sheet_name=SHEET_NAME,
        file_path=template_file_path,
        column_widths=[20] * len(fields),
        skip_rows=skip_rows,
        write_0=1,
    )

    add_images(
        images,
        workbook=wb,
        worksheet=SHEET_NAME,
        image_col=chr(65 + photo_index),
        skip_rows=skip_rows,
    )
    wb.active = wb[SHEET_NAME]

    def _get_column_range(fieldname):
        if fieldname in fields:
            col = get_column_letter(fields.index(fieldname) + 1)
            return "{}{}:{}{}".format(col, skip_rows + 1, col, len(excel_rows))

    cell_range = _get_column_range("packing_type")
    if cell_range:
        data_list = frappe.get_all("Packing Type Art", pluck="name")
        add_data_validation(wb, SHEET_NAME, data_list, cell_range)

    data_list = frappe.get_all("Matiere", pluck="name")
    for field in ["product_material1", "product_material2", "product_material3"]:
        cell_range = _get_column_range(field)
        if cell_range:
            add_data_validation(wb, SHEET_NAME, data_list, cell_range)

    # set supplier details
    if template == "lead_items_supplier_template" and supplier:
        wb[SHEET_NAME]["R1"] = re.sub(
            "<br>",
            "\n\n",
            frappe.render_template(
                SUPPLIER_DISPLAY_TEMPLATE,
                get_party_details(supplier, party_type="Supplier"),
            ),
        )
        wb[SHEET_NAME]["R1"].alignment = Alignment(horizontal="left")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def add_data_validation(wb, SHEET_NAME, data_list, cell_range):
    # Create a data-validation object with list validation

    list_sheet_name = frappe.generate_hash("", 6)

    sheet = wb.create_sheet(list_sheet_name)
    sheet.sheet_state = "hidden"
    for i in range(len(data_list)):
        sheet.cell(column=1, row=i + 1, value=data_list[i])

    # option 1: add values as range reference, create sheet and add list of values
    formula1 = "={}!$A$1:$A${}".format(list_sheet_name, len(data_list) + 1)

    # option 2: add values as comma seperated string.
    # formula1='"{}"'.format(",".join(data_list)),

    data_validation = DataValidation(type="list", allow_blank=True, formula1=formula1)

    data_validation.ranges.add(cell_range)
    wb[SHEET_NAME].add_data_validation(data_validation)


LEAD_ITEM_CONDITIONS = {
    "supplier_quotation": " and (is_disabled = 0 and is_quoted = 0)",
    "supplier_sample_request": " and (is_disabled = 0 and is_quoted = 1 and is_sample_validated = 0)",
    "create_lead_items": " and (is_disabled = 0 and is_sample_validated = 1)",
    "artyfetes": "",
}

SUPPLIER_DISPLAY_TEMPLATE = """
{{supplier_name}}
{{address_display}}
{{contact_person}}
{{contact_mobile}}
{{contact_email}}
"""
