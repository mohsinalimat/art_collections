# -*- coding: utf-8 -*-
# Copyright (c) 2019, GreyCube Technologies and contributors
# For license information, please see license.txt

from __future__ import unicode_literals

import re
from io import BytesIO
import io
import openpyxl
import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from frappe.utils import cint, get_site_url, get_url
import frappe

ILLEGAL_CHARACTERS_RE = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")

COL_MAP = {
    "your_ref": "Your Ref",
    "our_ref": "Our Ref",
    "excel_designation_cf": "designation",
    "description_1_cf": "Description 1",
    "description_2_cf": "Description 2",
    "description_3_cf": "Description 3",
    "other_language_cf": "Other Language",
    "genco": "GENCO",
    "nb_selling_packs_in_inner_art": "INNER",
    "packaging_description_cf": "Packaging Description",
}


def on_submit_purchase_order(doc, method=None):
    make_excel(doc.name, doc.doctype)


def on_submit_sales_order(doc, method=None):
    make_excel(doc.name, doc.doctype)


@frappe.whitelist()
def make_excel(docname=None, doctype=None):
    data = frappe.db.sql(
        """
        select 
            i.is_existing_product_cf,
            {your_ref} your_ref,
            i.item_code our_ref,
            i.excel_designation_cf,
            i.description_1_cf,
            i.description_2_cf,
            i.description_3_cf,
            i.other_language_cf,
            (select barcode from `tabItem Barcode` x where x.parent = i.name order by idx limit 1) genco,
            i.nb_selling_packs_in_inner_art,
            i.packaging_description_cf
        from 
        `tab{doctype}` dt
        inner join `tab{doctype} Item` dti on dti.parent = dt.name
        inner join tabItem i on i.name = dti.item_code
        where dt.name = %s
    """.format(
            your_ref=("dti.supplier_part_no" if doctype == "Purchase Order" else "''"),
            doctype=doctype,
        ),
        (docname,),
        as_dict=True,
        # debug=True,
    )

    existing_art_works = frappe.db.sql(
        """
    select 
        i.item_code,
        group_concat(concat(epaw.art_work_name,'||',epaw.art_work_attachment)) art_work
        from `tab{doctype}` dt
        inner join `tab{doctype} Item` dti on dti.parent = dt.name
        inner join tabItem i on i.name = dti.item_code
        inner join `tabExisting Product Art Work` epaw on epaw.parent = i.name
        where dt.name = %s
    group by i.name
    """.format(
            doctype=doctype
        ),
        (docname,),
        as_dict=True,
        # debug=True,
    )

    columns = [COL_MAP.get(col) for col, _ in data[0].items()][1:]

    site_url = get_url()

    # set the existing artworks for each item
    art_work_columns = []
    for d in existing_art_works:
        for row in filter(lambda x: x.our_ref == d.item_code, data):
            for tup in d.art_work.split(","):
                parts = tup.split("||")
                row[parts[0]] = f"{site_url}{parts[1]}"
                if not parts[0] in art_work_columns:
                    art_work_columns.append(parts[0])

    column_widths = [20, 20, 30, 30, 30, 20, 20, 15, 15, 30]

    wb = openpyxl.Workbook()
    # new
    product = [columns] + [
        list(d.values())[1:] for d in data if not cint(d.is_existing_product_cf)
    ]
    write_xlsx(product, "New Product", wb, column_widths)

    # existing
    product = [columns + art_work_columns] + [
        list(d.values())[1:] for d in data if cint(d.is_existing_product_cf)
    ]
    write_xlsx(product, "Existing Product", wb, column_widths)

    # make attachment
    out = io.BytesIO()
    wb.save(out)
    _file = frappe.get_doc(
        {
            "doctype": "File",
            "file_name": "{}.xlsx".format(docname),
            "attached_to_doctype": doctype,
            "attached_to_name": docname,
            "is_private": 1,
            "content": out.getvalue(),
        }
    )
    _file.save()
    frappe.db.commit()
    frappe.publish_realtime(
        "show_sales_order_email_dialog", {"user": frappe.session.user}
    )


def write_xlsx(data, sheet_name, wb=None, column_widths=None, file_path=None):
    # from xlsx utils with changes
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook(write_only=True)

    ws = wb.create_sheet(sheet_name, 0)

    for i, column_width in enumerate(column_widths):
        if column_width:
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width

    row1 = ws.row_dimensions[1]
    row1.font = Font(name="Calibri", bold=True)

    for idx, row in enumerate(data):
        clean_row = []
        print(row)
        for col, item in enumerate(row):
            value = item
            if isinstance(item, str) and next(
                ILLEGAL_CHARACTERS_RE.finditer(value), None
            ):
                # Remove illegal characters from the string
                value = re.sub(ILLEGAL_CHARACTERS_RE, "", value)

            if value:
                if isinstance(value, str) and value.startswith("http"):
                    _ = ws.cell(
                        column=col + 1, row=idx + 1, value=value.rsplit("/")[-1]
                    )
                    _.hyperlink = value
                else:
                    _ = ws.cell(column=col + 1, row=idx + 1, value=value)


def sample_write_xlsx(data, sheet_name, wb=None, column_widths=None, file_path=None):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "range names"
    for row in range(1, 40):
        ws1.append(range(600))
    ws2 = wb.create_sheet(title="Pi")
    ws2["F5"] = 3.14
    ws3 = wb.create_sheet(title="Data")
    for row in range(10, 20):
        for col in range(27, 54):
            _ = ws3.cell(
                column=col, row=row, value="{0}".format(get_column_letter(col))
            )
    print(ws3["AA10"].value)
    wb.save(filename="dest_filename")


# sales order print format
def get_print_context_for_art_collectons_sales_order(name):
    doc = frappe.get_doc("Sales Order", name)

    ctx = {"doc": {}}

    ctx["items"] = list(
        frappe.db.sql(
            """
        select i.item_name, i.customer_code, tib.barcode, i.customs_tariff_number,
        tw.warehouse_name, soi.price_list_rate, soi.total_saleable_qty_cf, 
        soi.net_rate, soi.net_amount, soi.description, soi.total_weight, 
        soi.qty, soi.image, so.overall_directive_art,
        if(soi.total_saleable_qty_cf >= soi.qty,1,0) in_stock
        from `tabSales Order Item` soi
        inner join `tabSales Order` so on so.name = soi.parent
        left outer join tabWarehouse tw on tw.name = soi.warehouse 
        inner join tabItem i on i.name = soi.item_code
        left outer join `tabItem Barcode` tib on tib.parent = i.name and tib.idx = 1 
        where soi.parent = %(name)s
    """,
            dict(name=name),
            as_dict=True,
        )
    )

    ctx["has_discount"] = any(x.discount_amount for x in doc.items)

    shipping_cost = 0
    taxes_cost = 0
    for tax in doc.taxes:
        account_type = frappe.db.get_value("Account", tax.account_head, "account_type")
        if account_type == "Tax":
            taxes_cost += tax.base_tax_amount
        else:
            shipping_cost += tax.base_tax_amount
    ctx["shipping_cost"] = shipping_cost
    ctx["taxes_cost"] = taxes_cost
    return ctx


# purchase order print format
def get_print_context_for_art_collectons_purchase_order(name):
    doc = frappe.get_doc("Purchase Order", name)

    ctx = {"doc": {}}

    ctx["items"] = list(
        frappe.db.sql(
            """
        select i.item_name, i.customer_code, tib.barcode, i.customs_tariff_number,
        tw.warehouse_name, poi.price_list_rate,  
        poi.net_rate, poi.net_amount, poi.description, poi.total_weight, 
        poi.qty, poi.image, po.overall_directive_art
        from `tabPurchase Order Item` poi
        inner join `tabPurchase Order` po on po.name = poi.parent
        left outer join tabWarehouse tw on tw.name = poi.warehouse 
        inner join tabItem i on i.name = poi.item_code
        left outer join `tabItem Barcode` tib on tib.parent = i.name and tib.idx = 1 
        where poi.parent = %(name)s
    """,
            dict(name=name),
            as_dict=True,
        )
    )

    ctx["has_discount"] = any(x.discount_amount for x in doc.items)

    shipping_cost = 0
    taxes_cost = 0
    for tax in doc.taxes:
        account_type = frappe.db.get_value("Account", tax.account_head, "account_type")
        if account_type == "Tax":
            taxes_cost += tax.base_tax_amount
        else:
            shipping_cost += tax.base_tax_amount
    ctx["shipping_cost"] = shipping_cost
    ctx["taxes_cost"] = taxes_cost

    print("*\n" * 10, ctx)
    return ctx


def get_so_excel_data(docname):
    items = frappe.db.sql(
        """
            select 
                i.item_name, i.customer_code, tib.barcode, i.customs_tariff_number,
                tw.warehouse_name, soi.price_list_rate, soi.total_saleable_qty_cf, 
                soi.net_rate, soi.net_amount, soi.discount_amount, soi.description, soi.total_weight, 
                soi.qty, soi.image, so.overall_directive_art,
                if(soi.total_saleable_qty_cf >= soi.qty,1,0) in_stock
            from `tabSales Order Item` soi
            inner join `tabSales Order` so on so.name = soi.parent
            left outer join tabWarehouse tw on tw.name = soi.warehouse 
            inner join tabItem i on i.name = soi.item_code
            left outer join `tabItem Barcode` tib on tib.parent = i.name and tib.idx = 1 
            where soi.parent = %s
    """,
        (docname),
        as_dict=True,
    )
    discontinued_items = frappe.db.sql(
        """
        select 
            i.item_name, i.customer_code, tib.barcode, i.customs_tariff_number,
            '' warehouse_name, 0 price_list_rate, 0 total_saleable_qty_cf, 
            0 net_rate, 0 net_amount, 0 discount_amount, soi.description, 0 total_weight, 
            soi.qty, '' image, so.overall_directive_art, 0 in_stock
        from `tabSales Order Discountinued Items CT` soi
        inner join `tabSales Order` so on so.name = soi.parent
        inner join tabItem i on i.name = soi.item_code
        left outer join `tabItem Barcode` tib on tib.parent = i.name and tib.idx = 1 
        where soi.parent = %s
    """,
        (docname),
        as_dict=True,
    )

    return {
        "in_stock_items": filter(lambda x: x.in_stock, items),
        "out_of_stock_items": filter(lambda x: not x.in_stock, items),
        "discontinued_items": discontinued_items,
    }


SO_COLUMNS = [
    ("warehouse_name", "Zone"),
    ("item_name", "Ref"),
    ("customer_code", "Your Ref. "),
    ("description", "Description"),
    ("barcode", "Barcode"),
    ("customs_tariff_number", "HScode"),
    ("weight", "Weight"),
    ("nb_selling_packs_in_inner_art", "Inner Qty"),
    ("qty", "Qty"),
    ("price_list_rate", "Gross Price"),
    ("discount_amount", "Discount"),
    ("net_rate", "Net Price"),
    ("net_amount", "Total Line"),
    ("image", "Photo"),
]


@frappe.whitelist()
def make_sales_order_excel(docname=None, doctype=None):
    all_items = get_so_excel_data(docname)
    columns = [[d[1] for d in SO_COLUMNS]]
    wb = openpyxl.Workbook()
    column_widths = [20, 20, 30, 30, 30, 20, 20, 15, 15, 30]

    data = [
        [d.get(col[0]) for d in all_items["discontinued_items"] for col in SO_COLUMNS]
    ]
    write_xlsx(columns + data, "Discontinued Items", wb, column_widths)

    data = [
        [d.get(col[0]) for d in all_items["out_of_stock_items"] for col in SO_COLUMNS]
    ]
    write_xlsx(columns + data, "Out of Stock Items", wb, column_widths)

    data = [[d.get(col[0]) for d in all_items["in_stock_items"] for col in SO_COLUMNS]]
    write_xlsx(columns + data, "In Stock Items", wb, column_widths)

    # make attachment
    out = io.BytesIO()
    wb.save(out)
    _file = frappe.get_doc(
        {
            "doctype": "File",
            "file_name": "{}.xlsx".format(docname),
            "attached_to_doctype": doctype,
            "attached_to_name": docname,
            "is_private": 1,
            "content": out.getvalue(),
        }
    )
    _file.save()
    frappe.db.commit()
    frappe.publish_realtime(
        "show_sales_order_email_dialog", {"user": frappe.session.user}
    )
