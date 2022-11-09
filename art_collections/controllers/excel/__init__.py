# -*- coding: utf-8 -*-
# Copyright (c) 2019, GreyCube Technologies and contributors
# For license information, please see license.txt

from __future__ import unicode_literals

import re
from io import BytesIO
import io
import openpyxl
import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from frappe.utils import cint, get_site_url, get_url, cstr, now_datetime
import frappe
import requests
import os

ILLEGAL_CHARACTERS_RE = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")


def write_xlsx(
    data,
    sheet_name,
    wb=None,
    column_widths=None,
    file_path=None,
    skip_rows=0,
    index=0,
    write_0=0,
):
    # from xlsx utils with changes
    column_widths = column_widths or []

    if file_path:
        wb = load_workbook(filename=file_path)
    elif wb is None:
        wb = openpyxl.Workbook()

    if not sheet_name in wb.sheetnames:
        wb.create_sheet(sheet_name, index)

    ws = wb.get_sheet_by_name(sheet_name)

    if not file_path:

        def _set_header():
            # set column widths
            for i, column_width in enumerate(column_widths):
                ws.column_dimensions[get_column_letter(i + 1)].width = (
                    column_width or 15
                )
            # set row style
            row1 = ws.row_dimensions[1]
            row1.font = Font(name="Calibri", bold=True)

        _set_header()

    for idx, row in enumerate(data):
        clean_row = []
        for col, item in enumerate(row):
            value = item
            if isinstance(item, str) and next(
                ILLEGAL_CHARACTERS_RE.finditer(value), None
            ):
                # Remove illegal characters from the string
                value = re.sub(ILLEGAL_CHARACTERS_RE, "", value)

            if value or write_0 and value == 0:
                if isinstance(value, str) and value.startswith("http"):
                    _ = ws.cell(
                        column=col + 1,
                        row=idx + skip_rows + 1,
                        value=value.rsplit("/")[-1],
                    )
                    _.hyperlink = value
                else:
                    _ = ws.cell(column=col + 1, row=idx + skip_rows + 1, value=value)
    return wb


def attach_file(content, **args):
    _file = frappe.get_doc(
        {
            "doctype": "File",
            "file_name": args.get("file_name")
            or get_with_now(args.get("docname"), ".xlsx"),
            "attached_to_doctype": args.get("doctype"),
            "attached_to_name": args.get("docname"),
            "is_private": 1,
            "content": content,
        }
    )
    _file.save()
    frappe.db.commit()
    if args.get("show_email_dialog"):
        frappe.publish_realtime("show_email_dialog", args, user=frappe.session.user)
    return _file


def get_with_now(filename, extn):
    return "{}_{}{}".format(filename, now_datetime().strftime("%y%m%d%H%M"), extn)


def add_images(data, workbook, worksheet="", image_col="S", skip_rows=0):
    ws = workbook.get_sheet_by_name(worksheet)
    for row, image_url in enumerate(data):
        if image_url:
            _filename, extension = os.path.splitext(image_url)
            if extension.lower() in [".png", ".jpg", ".jpeg"]:
                try:
                    content = None

                    if image_url.startswith("http"):
                        content = requests.get(image_url).content
                    else:
                        item_file = frappe.get_doc("File", {"file_url": image_url})
                        content = item_file.get_content()
                    if content:
                        image = openpyxl.drawing.image.Image(io.BytesIO(content))
                        image.height = 100
                        image.width = 100
                        ws.add_image(image, f"{image_col}{cstr(row+1 + skip_rows)}")
                        ws.row_dimensions[row + 1 + skip_rows].height = 90
                except Exception as e:
                    print(e)
                    pass


def sample_write_xlsx():
    """
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "range names"
    for row in range(1, 40):
        ws1.append(range(600))
    ws2 = wb.create_sheet(title="Pi")
    ws2["F5"] = 3.14
    ws3 = wb.create_sheet(title="Data")
    for row in range(10, 20):
        for col in range(27, 54):
            _ = ws3.cell(
                column=col, row=row, value="{0}".format(get_column_letter(col))
            )
    wb.save(filename="dest_filename")
    """
    pass
