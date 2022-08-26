from __future__ import unicode_literals
from locale import currency
from unittest import skip
import frappe
from frappe import _
import io
import openpyxl
from frappe.utils import cint, get_site_url, get_url
from art_collections.controllers.excel import write_xlsx, attach_file, add_images
from openpyxl import load_workbook, Workbook
from art_collections.art_collections.doctype.sales_confirmation.sales_confirmation import (
    make_from_po,
)
import os


def on_submit_purchase_order(doc, method=None):
    _make_excel_attachment(doc.doctype, doc.name)

    # set is_po_created 1 in lead items connected to PO Items
    frappe.db.sql(
        """
    update `tabPurchase Order Item` tpoi 
	inner join tabItem ti on ti.item_code = tpoi.item_code 
	inner join `tabLead Item` tli on tli.name = ti.lead_item_cf 
	set tli.is_po_created = 1 
	where tpoi.parent = %s
    """,
        (doc.name,),
    )


@frappe.whitelist()
def _make_excel_attachment(doctype, docname):
    """for product excel"""

    data, existing_art_works = [], []
    site_url = get_url()

    data = frappe.db.sql(
        """
        select 
            i.item_code, i.item_name ,
            poi.supplier_part_no ,
            tib.barcode,
            i.customs_tariff_number ,
            poi.qty,
            poi.stock_uom , 
            poi.rate , 
			poi.amount ,
            i.is_existing_product_cf ,
            case when i.image is null then ''
                when SUBSTR(i.image,1,4) = 'http' then i.image
                else concat('{}',i.image) end image ,
            i.image image_url    
        from `tabPurchase Order Item` poi
        inner join `tabPurchase Order` po on po.name = poi.parent
        inner join tabItem i on i.name = poi.item_code
        left outer join `tabItem Barcode` tib on tib.parent = i.name 
            and tib.idx  = (
                select min(idx) from `tabItem Barcode` tib2
                where parent = i.name
            )
        where poi.parent = %s
    """.format(
            get_url()
        ),
        (docname,),
        as_dict=True,
    )

    currency = frappe.db.get_value(doctype, docname, "currency")

    columns = [
        _("Item Code"),
        _("Item Name"),
        _("Supplier items"),
        _("Barcode"),
        _("HSCode"),
        _("Quantity"),
        _("Stock UOM"),
        _("Rate ({0})").format(currency),
        _("Amount ({0})").format(currency),
        _("Photo"),
    ]

    fields = [
        "item_code",
        "item_name",
        "supplier_part_no",
        "barcode",
        "customs_tariff_number",
        "qty",
        "stock_uom",
        "rate",
        "amount",
        "image",
    ]

    wb = openpyxl.Workbook()

    # new art works
    excel_rows, images = [columns], [""]
    for d in data:
        if not cint(d.is_existing_product_cf):
            excel_rows.append([d.get(f) for f in fields])
            images.append(d.get("image_url"))
    write_xlsx(excel_rows, "New Product", wb, [20] * len(columns))
    add_images(images, workbook=wb, worksheet="New Product", image_col="K")

    # existing art works
    art_works = frappe.db.sql(
        """
    select 
        DISTINCT parent item_code, art_work_name , art_work_attachment
    from `tabExisting Product Art Work`
    where parent in (
    	select item_code from `tabPurchase Order Item` tpoi
    	where parent = %s
    )
    """,
        (docname,),
        as_dict=True,
    )
    art_work_names = frappe.utils.unique([d.art_work_name for d in art_works])
    excel_rows, images = [columns + art_work_names], [""]
    # print(art_work_names, art_works, excel_rows)
    for d in data:
        if cint(d.is_existing_product_cf):
            excel_rows.append([d.get(f) for f in fields])
            images.append(d.get("image_url"))
            for name in art_work_names:
                for aw in art_works:
                    if aw.item_code == d.item_code and aw.art_work_name == name:
                        excel_rows[-1].append(f"{site_url}{aw.art_work_attachment}")
    write_xlsx(excel_rows, "Existing Product", wb, [20] * len(excel_rows[0]))
    add_images(images, workbook=wb, worksheet="New Product", image_col="M")

    # make attachment
    out = io.BytesIO()
    wb.save(out)
    attach_file(
        out.getvalue(),
        doctype=doctype,
        docname=docname,
        show_email_dialog=1,
    )


@frappe.whitelist()
def supplier_email_callback(docname):
    frappe.db.set_value("Purchase Order", docname, "replied_to_supplier_cf", 1)
    # create/update sales confirmation
    sc_name = make_from_po(docname)
    return sc_name


@frappe.whitelist()
def make_supplier_email_attachments(po_name):
    """
    1. item details excel same as Item Import template used in Data Import for items
    2. Packaging Description.
    3. PO Pdf
    """

    # item details excel
    data = []
    from art_collections.controllers.item_import import get_records

    if frappe.db.exists("Purchase Order", po_name):
        filter_sql = """
        where exists (select 1 from `tabPurchase Order Item` x where x.item_code = ti.item_code and x.parent = '{}')
        """.format(
            po_name
        )
        kwargs = {"filter_sql": filter_sql, "doctype": "Item", "as_dict": True}
        data = get_records(**kwargs)
        print([d for d in data[0].keys()])

    template_file_path = os.path.join(
        os.path.dirname(__file__),
        "purchase_order_supplier_email_item_details.xlsx",
    )
    workbook = load_workbook(template_file_path)

    sheet = workbook["configuration"]
    fields = [col.value for col in sheet["B"][1:]]
    skip_rows = cint(sheet["C2"].value)

    SHEET_NAME = "Item Details"
    sheet = workbook[SHEET_NAME]
    sheet.delete_rows(skip_rows + 1, sheet.max_row - 1)
    workbook.active = workbook[SHEET_NAME]

    # take fields from configuration in excel
    data = [[d.get(col) for col in fields] for d in data]

    for row, d in enumerate(data):
        for col, value in enumerate(d):
            sheet.cell(row=row + skip_rows + 1, column=col + 1, value=value)
    out = io.BytesIO()
    workbook.save(out)

    attach_file(
        out.getvalue(),
        doctype="Purchase Order",
        docname=po_name,
        file_name="Supplier Item Details %s.xlsx" % (po_name),
        show_email_dialog=0,
    )

    # packaging description
    data = frappe.db.sql(
        """
        select 
        ti.customer_code , ti.item_code , ti.excel_designation_cf ,
        ti.description_1_cf , ti.description_2_cf , ti.description_3_cf , 
        other_language_cf , tib.barcode , nb_inner_in_outer_art , 
        ti.description 
        from tabItem ti 
        inner join `tabPurchase Order Item` tpoi on tpoi.item_code = ti.item_code and tpoi.parent = %s
        left outer join `tabItem Barcode` tib on tib.parent = ti.name  and tib.barcode_type = 'EAN'
    """,
        (po_name,),
        as_list=True,
    )

    data = [PACKAGING_DESCRIPTION_HEADER] + data

    wb = write_xlsx(
        data,
        sheet_name="Sales Confirmation Details",
        column_widths=[20] * len(PACKAGING_DESCRIPTION_HEADER),
        skip_rows=0,
    )
    out = io.BytesIO()
    wb.save(out)
    attach_file(
        out.getvalue(),
        doctype="Purchase Order",
        docname=po_name,
        file_name="Supplier Packaging Description %s.xlsx" % (po_name),
        show_email_dialog=0,
    )

    # PO Pdf

    auto_formats = (
        frappe.db.get_single_value(
            "Art Collections Settings", "art_auto_email_template"
        )
        or []
    )

    print_format = [d.doc_type for d in auto_formats]

    out = frappe.attach_print(
        "Purchase Order",
        po_name,
        print_format=print_format and print_format[0],
    )

    attach_file(
        out["fcontent"],
        doctype="Purchase Order",
        docname=po_name,
        file_name="Purchase Order {} {}.pdf".format(po_name, frappe.utils.today()),
        show_email_dialog=1,
        callback="supplier_email_callback",
    )


PACKAGING_DESCRIPTION_HEADER = [
    "Your ref",
    "Our Ref",
    "Designation",
    "DESCRIPTION 1",
    "DESCRIPTION 2",
    "DESCRIPTION 3",
    "OTHER LANGUAGES ",
    "GENCOD",
    "INNER ",
    "Description",
]
