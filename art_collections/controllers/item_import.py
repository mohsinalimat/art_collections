# -*- coding: utf-8 -*-
import frappe
from frappe import _
from frappe.utils import cstr
from frappe.utils.csvutils import to_csv
from openpyxl import load_workbook, Workbook
from frappe.utils.xlsxutils import make_xlsx
from frappe.utils import today, get_site_url
from frappe.utils.xlsxutils import build_xlsx_response
import io
import os

frappe_template_columns = [
    "Item Code",
    "Item Name",
    "Item Group",
    "Default Unit of Measure",
    "Qty In Selling Pack",
    "Thickness (cm)",
    "Width (cm)",
    "Length (cm)",
    "Weight (kg)",
    "HS CODE",
    "Minimum Order Qty",
    "Main Design Color",
    "Packing Type",
    "Supplier (Supplier Items)",
    "Supplier Part Number (Supplier Items)",
    "ID (Item Components)",
    "Matiere (Item Components)",
    "ID (Barcodes)",
    "Barcode (Barcodes)",
    "ID (UOMs)",
    "UOM (UOMs)",
    "Conversion Factor (UOMs)",
    "ID (Product Packing Dimensions)",
    "UOM (Product Packing Dimensions)",
    "Materials (Product Packing Dimensions)",
    "Length (cm) (Product Packing Dimensions)",
    "Width (cm) (Product Packing Dimensions)",
    "Thickness (cm) (Product Packing Dimensions)",
    "Weight (kg) (Product Packing Dimensions)",
    "CBM (m3) (Product Packing Dimensions)",
]

custom_template_columns = [
    "Item Code",
    "Item Name",
    "Item Group",
    "Supplier (Supplier Items)",
    "Supplier Part Number (Supplier Items)",
    "Minimum Order Qty",
    "HS CODE",
    "Packing Type",
    "Default Unit of Measure",
    "Length (cm)",
    "Width (cm)",
    "Thickness (cm)",
    "Weight (kg)",
    "Main Design Color",
    "ID (Item Components)",
    "Matiere (Item Components)",
    "ID (Barcodes)",
    "Barcode (Barcodes)",
    "ID (UOMs)",
    "UOM (UOMs)",
    "Conversion Factor (UOMs)",
    "ID (Product Packing Dimensions)",
    "Length (cm) (Product Packing Dimensions)",
    "Width (cm) (Product Packing Dimensions)",
    "Thickness (cm) (Product Packing Dimensions)",
    "Weight (kg) (Product Packing Dimensions)",
    "CBM (m3) (Product Packing Dimensions)",
    "Materials (Product Packing Dimensions)",
    "ID (UOMs)",
    "UOM (UOMs)",
    "Conversion Factor (UOMs)",
    "ID (Product Packing Dimensions)",
    "Length (cm) (Product Packing Dimensions)",
    "Width (cm) (Product Packing Dimensions)",
    "Thickness (cm) (Product Packing Dimensions)",
    "Weight (kg) (Product Packing Dimensions)",
    "CBM (m3) (Product Packing Dimensions)",
    "Materials (Product Packing Dimensions)",
    "ID (UOMs)",
    "UOM (UOMs)",
    "Conversion Factor (UOMs)",
    "ID (Product Packing Dimensions)",
    "Length (cm) (Product Packing Dimensions)",
    "Width (cm) (Product Packing Dimensions)",
    "Thickness (cm) (Product Packing Dimensions)",
    "Weight (kg) (Product Packing Dimensions)",
    "CBM (m3) (Product Packing Dimensions)",
    "Materials (Product Packing Dimensions)",
    "ID (UOMs)",
    "UOM (UOMs)",
    "Conversion Factor (UOMs)",
    "ID (Product Packing Dimensions)",
    "Length (cm) (Product Packing Dimensions)",
    "Width (cm) (Product Packing Dimensions)",
    "Thickness (cm) (Product Packing Dimensions)",
    "Weight (kg) (Product Packing Dimensions)",
    "CBM (m3) (Product Packing Dimensions)",
    "Materials (Product Packing Dimensions)",
]


@frappe.whitelist()
def download_template(data_import=None, **kwargs):
    template_path = frappe.get_app_path(
        "art_collections", "controllers", "item_import_template.xlsx"
    )

    template = load_workbook(template_path)
    sheet = template["Sheet1"]
    sheet.delete_rows(3, sheet.max_row - 1)

    data = get_records(**kwargs)

    for row, d in enumerate(data):
        for col, value in enumerate(d):
            sheet.cell(row=row + 3, column=col + 1, value=value)
    out = io.BytesIO()
    template.save(out)
    frappe.response["filename"] = _("Item") + ".xlsx"
    frappe.response["filecontent"] = out.getvalue()
    frappe.response["type"] = "binary"


def start_item_import(doc, method):
    # return
    # transform raw import file to frappe format and attach to doc

    if doc.reference_doctype == "Item" and doc.import_art_item_file:
        file_name = doc.import_art_item_file.split("/")[-1]
        file_path = frappe.get_site_path("private", "files", file_name)
        # file_path = "/home/frappe/frappe-13/sites/debug/item_import.xlsx"

        wb = load_workbook(filename=file_path)

        def get_values(col, row):
            # Copy UOM (UOMs) to UOM (Product Packing Dimensions)
            # to save repeating UOM values for PPD child table in upload excel
            if col == "UOM (Product Packing Dimensions)":
                col = "UOM (UOMs)"
            value = [
                row[idx] for idx, d in enumerate(custom_template_columns) if d == col
            ]
            if col in ["HS CODE"]:
                return [cstr(x) for x in value]
            return value

        docs_to_create, items = [], []

        def check_and_create(row, column, doctype, field="name"):
            print(row, type(row))
            values = get_values(column, row)
            if values and values[0]:
                name = frappe.db.sql(
                    """select name from `tab{}` where name = %s limit 1""".format(
                        doctype
                    ),
                    (values[0],),
                )
                if not name:
                    frappe.get_doc({"doctype": doctype, field: values[0]}).insert()
                    # handle case when values are in wrong case:
                    # frappe does not allow creating doc with name in different case
                elif not name[0][0] == values[0]:
                    row[custom_template_columns.index(column)] = name[0][0]

            return row

        for row in wb["Sheet1"].iter_rows(
            min_row=3, max_row=100, max_col=100, values_only=True
        ):
            if not row[0]:
                break

            row = list(row)

            # create docs (hscode, design color, matiere, packing type art) that do not exist
            row = check_and_create(
                row, "HS CODE", "Customs Tariff Number", "tariff_number"
            )

            row = check_and_create(
                row, "Main Design Color", "Design Color", "design_color"
            )

            row = check_and_create(
                row, "Matiere (Item Components)", "Matiere", "matiere"
            )

            row = check_and_create(
                row, "Packing Type", "Packing Type Art", "packing_type"
            )

            tmp = [get_values(col, row) for col in frappe_template_columns]

            qty_in_selling_pack_art = row[
                custom_template_columns.index("Conversion Factor (UOMs)")
            ]

            tmp[frappe_template_columns.index("Qty In Selling Pack")] = [
                qty_in_selling_pack_art
            ]
            items.append(tmp)

        import_csv = [frappe_template_columns]
        for item in items:
            max_child_count = max([len(col) for col in item])
            for i in range(len(item)):
                item[i] = item[i] + ([""] * (max_child_count - len(item[i])))
            for idx in range(max_child_count):
                import_csv.append([col[idx] for col in item])
            # print(item)

        # remove blank rows
        import_csv = [d for d in import_csv if not set(d) == [""]]

        xlsx_file = make_xlsx(import_csv, "Data Import Template")
        file_data = xlsx_file.getvalue()

        with open("sample_import_test.xlsx", "wb") as item_import_file:
            item_import_file.write(file_data)

        f = frappe.get_doc(
            doctype="File",
            content=file_data,
            file_name="item_import %s.xlsx" % frappe.generate_hash("", 6),
            is_private=1,
        )

        f.save(ignore_permissions=True)
        doc.import_file = f.file_url


def get_records(**kwargs):
    conditions, limit = "", ""

    if kwargs.get("export_records") == "5_records":
        limit = " limit 5 "
    elif kwargs.get("export_records") == "blank_template":
        limit = " limit 0 "
    elif kwargs.get("export_records") == "by_filter":
        export_filters = frappe.parse_json(kwargs.get("export_filters"))
        parent_data = frappe.db.get_list(
            kwargs.get("doctype"),
            filters=export_filters,
            fields=["name"],
            as_list=1,
        )

        if parent_data:
            conditions = " where ti.name in (%s)" % ",".join(
                ["'%s'" % d[0] for d in parent_data]
            )
    elif kwargs.get("filter_sql"):
        conditions = kwargs.get("filter_sql")

    return frappe.db.sql(
        """
with selling_pack as
(
    select 
        ti.name, 
        tucd.name id_uoms_selling_pack ,
        tucd.uom uom_uoms_selling_pack ,
        tucd.conversion_factor conversion_factor_uoms_selling_pack ,
        tppd.name id_ppd_selling_pack , 
        tppd.`length` length_ppd_selling_pack ,
        tppd.thickness thickness_ppd_selling_pack ,
        tppd.materials materials_ppd_selling_pack ,
        tppd.uom uom_ppd_selling_pack ,
        tppd.width width_ppd_selling_pack ,
        tppd.weight weight_ppd_selling_pack ,
        tppd.cbm cbm_ppd_selling_pack 
    from tabItem ti
    left join `tabProduct Packing Dimensions` tppd on tppd.parent = ti.name and tppd.uom = 'Selling Pack'
    left outer join `tabUOM Conversion Detail` tucd on tucd.parent = ti.name and tucd.uom = 'Selling Pack'
),
inner_carton as
(
    select 
        ti.name, 
        tucd.name id_uoms_inner_carton ,
        tucd.uom uom_uoms_inner_carton ,
        tucd.conversion_factor conversion_factor_uoms_inner_carton ,
        tppd.name id_ppd_inner_carton , 
        tppd.`length` length_ppd_inner_carton ,
        tppd.thickness thickness_ppd_inner_carton ,
        tppd.materials materials_ppd_inner_carton ,
        tppd.uom uom_ppd_inner_carton ,
        tppd.width width_ppd_inner_carton ,
        tppd.weight weight_ppd_inner_carton ,		
        tppd.cbm cbm_ppd_inner_carton 
    from tabItem ti
    left join `tabProduct Packing Dimensions` tppd on tppd.parent = ti.name and tppd.uom = 'Inner Carton'
    left outer join `tabUOM Conversion Detail` tucd on tucd.parent = ti.name and tucd.uom = 'Inner Carton'
),
maxi_inner as
(
    select 
        ti.name, 
        tucd.name id_uoms_maxi_inner ,
        tucd.uom uom_uoms_maxi_inner ,
        tucd.conversion_factor conversion_factor_uoms_maxi_inner ,
        tppd.name id_ppd_maxi_inner , 
        tppd.`length` length_ppd_maxi_inner ,
        tppd.thickness thickness_ppd_maxi_inner ,
        tppd.materials materials_ppd_maxi_inner ,
        tppd.uom uom_ppd_maxi_inner ,
        tppd.width width_ppd_maxi_inner ,
        tppd.weight weight_ppd_maxi_inner ,		
        tppd.cbm cbm_ppd_maxi_inner 
    from tabItem ti
    left join `tabProduct Packing Dimensions` tppd on tppd.parent = ti.name and tppd.uom = 'Maxi Inner'
    left outer join `tabUOM Conversion Detail` tucd on tucd.parent = ti.name and tucd.uom = 'Maxi Inner'
),
outer_carton as
(
    select 
        ti.name, 
        tucd.name id_uoms_outer_carton ,
        tucd.uom uom_uoms_outer_carton ,
        tucd.conversion_factor conversion_factor_uoms_outer_carton ,
        tppd.name id_ppd_outer_carton , 
        tppd.`length` length_ppd_outer_carton ,
        tppd.thickness thickness_ppd_outer_carton ,
        tppd.materials materials_ppd_outer_carton ,
        tppd.uom uom_ppd_outer_carton ,
        tppd.width width_ppd_outer_carton ,
        tppd.weight weight_ppd_outer_carton ,		
        tppd.cbm cbm_ppd_outer_carton 
    from tabItem ti
    left join `tabProduct Packing Dimensions` tppd on tppd.parent = ti.name and tppd.uom = 'Outer Carton'
    left outer join `tabUOM Conversion Detail` tucd on tucd.parent = ti.name and tucd.uom = 'Outer Carton'
)
select ti.item_code, ti.item_name , ti.item_group ,
tis.supplier , tis.supplier_part_no , 
ti.min_order_qty , ti.customs_tariff_number , ti.packing_type_art , ti.stock_uom , 
ti.length_art , ti.width_art , ti.thickness_art , ti.weight_art , ti.main_design_color_art , 
tica.name id_item_components , tica.matiere matiere_item_components ,
tib.name id_barcodes, tib.barcode barcode_barcodes , 
sel.id_uoms_selling_pack , 
sel.uom_uoms_selling_pack , 
sel.conversion_factor_uoms_selling_pack ,
sel.id_ppd_selling_pack , 
sel.length_ppd_selling_pack ,
sel.width_ppd_selling_pack ,
sel.thickness_ppd_selling_pack ,
sel.weight_ppd_selling_pack ,
sel.cbm_ppd_selling_pack ,
sel.materials_ppd_selling_pack ,
-- sel.uom_ppd_selling_pack ,
ic.id_uoms_inner_carton , 
ic.uom_uoms_inner_carton , 
ic.conversion_factor_uoms_inner_carton ,
ic.id_ppd_inner_carton , 
ic.length_ppd_inner_carton ,
ic.width_ppd_inner_carton ,
ic.thickness_ppd_inner_carton ,
ic.weight_ppd_inner_carton ,
ic.cbm_ppd_inner_carton ,
ic.materials_ppd_inner_carton ,
-- ic.uom_ppd_inner_carton ,
mi.id_uoms_maxi_inner , 
mi.uom_uoms_maxi_inner , 
mi.conversion_factor_uoms_maxi_inner ,
mi.id_ppd_maxi_inner , 
mi.length_ppd_maxi_inner ,
mi.width_ppd_maxi_inner ,
mi.thickness_ppd_maxi_inner ,
mi.weight_ppd_maxi_inner ,
mi.cbm_ppd_maxi_inner ,
mi.materials_ppd_maxi_inner ,
-- mi.uom_ppd_maxi_inner ,
oc.id_uoms_outer_carton , 
oc.uom_uoms_outer_carton , 
oc.conversion_factor_uoms_outer_carton ,
oc.id_ppd_outer_carton , 
oc.length_ppd_outer_carton ,
oc.width_ppd_outer_carton ,
oc.thickness_ppd_outer_carton ,
oc.weight_ppd_outer_carton ,
oc.cbm_ppd_outer_carton ,
oc.materials_ppd_outer_carton
-- oc.uom_ppd_outer_carton ,
from tabItem ti 
left outer join `tabItem Supplier` tis on tis.parent = ti.name 
left outer join (select name, matiere , parent from `tabItem Components Art` group by parent) tica on tica.parent = ti.name
left outer join (select name , barcode , parent from `tabItem Barcode` group by parent) tib on tib.parent = ti.name
left outer join selling_pack sel on sel.name = ti.name
left outer join inner_carton ic on ic.name = ti.name
left outer join maxi_inner mi on mi.name = ti.name
left outer join outer_carton oc on oc.name = ti.name
{conditions} {limit}
    """.format(
            conditions=conditions, limit=limit
        )
    )
